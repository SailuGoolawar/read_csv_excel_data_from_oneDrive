# ----------------------------------------------------------------------------------------------------------------------
# Script        : oneDrive.py
# Author        : Sailu Goolawar
# Created On    : 23/02/2021
# Dependencies  : requests,base64,StringIo,BytesIO from io,csv,openpyxl
# Description   : this script will useful to read the csv/excel data from the oneDrive Shared link
# Last Modified : 23/02/2021
# ----------------------------------------------------------------------------------------------------------------------
import requests
import base64
from io import StringIO, BytesIO
import csv
import openpyxl

# please provide your files shared link here
file_link = "Your Shared Link Here "
# please provide the file types as csv or excel
file_type = "excel"
# if the file_type is excel please provide the sheetName, if don't we will take the first sheet.
sheetName = 'Sheet1'


# this function will generate direct download link for the shared oneDrive file link
def get_onedrive_directdownloadlink (onedrive_file_link):
    # step one : encoding the url into base64 format
    data_bytes64 = base64.b64encode(bytes(onedrive_file_link, 'utf-8'))
    # step two : making unpadded base64url format by replacing '/' with '_' and '+' with '-' and removing '=' from the end
    data_bytes64_String = data_bytes64.decode('utf-8').replace('/','_').replace('+','-').rstrip("=")
    # step three : appending 'u!' to the beginning of the encoded url
    resultUrl = f"https://api.onedrive.com/v1.0/shares/u!{data_bytes64_String}/root/content"
    # returning direct_download link
    return resultUrl


# error text style
def error(text):
    return f"\033[{'31m'}{text}\033[00m"


# warning text style
def warning(text):
    return f"\033[{'93m'}{text}\033[00m"


# below function will reads the csv data from the url and returns the data
def get_csv_data(data_link):
    # getting the csv content from the file using requests.get() method
    url = requests.get(data_link).text
    # getting csv raw data from the url
    csv_raw = StringIO(url)
    # reading the csv using csv.reader()
    read = csv.reader(csv_raw)
    # empty csv_data
    csv_data = []
    # looping the csv data and storing it into the csv_data list
    for d in read:
        csv_data.append(d)
    # returning the csv_data
    return csv_data


def get_excel_data(data_link,sheet_name):
    # getting the excel content from the file using requests.get() method
    url = requests.get(data_link).content
    csv_data_from_excel = []
    try:
        # opening the workbook using openpyxl
        wb = openpyxl.load_workbook(BytesIO(url))
        # getting sheets in the workbook
        sheets = wb.sheetnames
        # checking if the given sheet is present in the workbook or not
        if sheet_name in sheets:
            # if sheet_name found getting the data and making a nested list with the rows.
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                csv_data_from_excel.append(list(row))
        else:
            # if sheet_name not found/ sheet_name not provided will take the first Sheet by default
            if sheet_name != '':
                print(warning(str(sheet_name)+' Not found. taking the first Sheet'))
            ws = wb[sheets[0]]
            for row in ws.iter_rows(values_only=True):
                csv_data_from_excel.append(list(row))
    except:
        # if any error occurs while getting the workbook from url error will display.
        print(error('Something went wrong, Please re upload the file link.'))
    return csv_data_from_excel


# calling get_onedrive_directdownloadlink() function to get the download link
response = get_onedrive_directdownloadlink(file_link)


if __name__ == "__main__":
    if file_type.lower() == 'csv':
        # --------------------------------------------------------------------------------------------------------------
        # in case of csv
        # --------------------------------------------------------------------------------------------------------------
        csv_data = get_csv_data(response)
        print(csv_data)
    elif file_type.lower() == 'excel':
        # --------------------------------------------------------------------------------------------------------------
        # in case of excel
        # --------------------------------------------------------------------------------------------------------------
        excel_data = get_excel_data(response, sheetName)
        print(excel_data)
    else:
        print(error('We accept csv and excel file types only.'))